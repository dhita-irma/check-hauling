import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import WHITE
from openpyxl.styles.colors import YELLOW
from openpyxl.styles.colors import RED


wb = xl.load_workbook('dataBB_RM.xlsx')
dataBB = wb.active

redFill = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
whiteFill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
yellowFill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')

print('Reading rows...')
for rowBB in range(7, dataBB.max_row + 1):
    current_voucherBB = int(dataBB.cell(rowBB, 4).value[:5])
    if dataBB.cell(rowBB+1, 4).value is not None:
        next_voucherBB = int(str(dataBB.cell(rowBB+1, 4).value)[:5])
        if next_voucherBB - current_voucherBB == 0:
            dataBB.cell(rowBB, 4).fill = redFill
            dataBB.cell(rowBB+1, 4).fill = redFill
        elif next_voucherBB - current_voucherBB > 1:
            dataBB.cell(rowBB, 4).fill = yellowFill
        else:
            dataBB.cell(rowBB, 4).fill = whiteFill


print('Saving workbook...')
wb.save('dataBB_RM_checked.xlsx')