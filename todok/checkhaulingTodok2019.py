#! python3
# check_hauling.py

import openpyxl as xl
wb1 = xl.load_workbook('Hauling Todok 2019.xlsx')
haulingRM = wb1.active

wb2 = xl.load_workbook('dataBB_todok.xlsx')
dataBB = wb2.active

print('Reading rows...')
for row1 in range(3, haulingRM.max_row + 1):
    voucher1 = haulingRM.cell(row=row1, column=2).value
    haulingRM.cell(row1, 1).value = 'PENDING'
    for row2 in range(7, dataBB.max_row + 1):
        voucher2 = int(str(dataBB.cell(row2, 4).value)[:5])
        if voucher2 == voucher1:
            for col in range(1, dataBB.max_column + 1):
                haulingRM.cell(row=row1, column=col + 2).value = dataBB.cell(row=row2, column=col).value
                haulingRM.cell(row=row1, column=1).value = 'OK'

print('Saving workbook...')
wb1.save('Hauling Todok 2019 update.xlsx')
