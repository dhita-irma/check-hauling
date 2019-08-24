#! python3
#  compareMasterBBM.py

import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW

# load two workbooks
wb1 = xl.load_workbook('BBM 0701D.xlsx')
sheet1 = wb1.active

wb2 = xl.load_workbook('MasterBBM 0701.xlsx')
sheet2 = wb2.active

yellowFill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')

print('Reading rows...')
for rowNum in range(5, sheet1.max_row + 1):
    value1 = sheet1.cell(row=rowNum, column=6).value
    value2 = sheet2.cell(row=rowNum, column=14).value
    if value1 != value2:
        sheet1.cell(row=rowNum, column=6).value = f"{str(value1)} >> {str(value2)}"
        sheet1.cell(row=rowNum, column=6).fill = yellowFill

print('Saving workbook...')
wb1.save('crosscheck_masterBBM.xlsx')